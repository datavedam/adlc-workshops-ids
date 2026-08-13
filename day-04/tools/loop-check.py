#!/usr/bin/env python3
"""
Prove the Day 4 loop ran, before anybody claims it did.

Four proofs, and each one is a path that has to open:

  1  the loop record       every stage named, with an output path
  2  the tracker           measured columns filled, no closed row without evidence
  3  the stop log          every stop marked decision or lookup
  4  the guard record      two refusals with exit 2, and one allowed write

It reports what is absent. It never repairs anything, because a tool that
repairs its own evidence proves nothing.

Run:  python3 tools/loop-check.py [--root .] [--file tracker.xlsx]
Exit code 1 when a proof is absent or thin.
"""

import argparse
import json
import os
import re
import sys

STAGES = ["understand", "context", "grill", "decide",
          "specify", "plan", "build", "prove"]
MEASURED = {"harness": 33, "model": 34, "tokens": 35, "rounds": 37}
EVIDENCE_COL, STATUS_COL, ID_COL, NAME_COL = 41, 11, 1, 4


class Report:
    def __init__(self):
        self.rows = []

    def add(self, proof, ok, detail):
        self.rows.append((proof, ok, detail))

    def done(self):
        width = max(len(p) for p, _, _ in self.rows)
        bad = 0
        for proof, ok, detail in self.rows:
            mark = "OK  " if ok else "GAP "
            bad += 0 if ok else 1
            print(f"{mark}{proof:<{width}}  {detail}")
        print()
        if bad:
            print(f"FAIL — {bad} of {len(self.rows)} proofs are absent or thin.")
            print("Repair them tonight. Day 5 reads these paths.")
            return 1
        print(f"PASS — all {len(self.rows)} proofs are present.")
        print("Day 5 is a review, not a scramble.")
        return 0


def check_loop_record(root, rep):
    path = os.path.join(root, "evidence", "day-04", "loop-run.md")
    if not os.path.exists(path):
        rep.add("1 loop record", False, f"absent: {path}")
        return
    text = open(path, encoding="utf-8").read().lower()
    absent = [s for s in STAGES if s not in text]
    # A stage named with no path beside it is a claim, not a record.
    paths = len(re.findall(r"[\w./-]+\.(?:md|json|txt|py|xlsx|log)", text))
    if absent:
        rep.add("1 loop record", False,
                f"{len(STAGES) - len(absent)}/8 stages named, absent: "
                + ", ".join(absent))
    elif paths < 8:
        rep.add("1 loop record", False,
                f"8 stages named but only {paths} output path(s). One for each stage.")
    else:
        rep.add("1 loop record", True, f"8 stages, {paths} output paths · {path}")


def check_tracker(root, book, rep):
    path = book if os.path.isabs(book) else os.path.join(root, book)
    if not os.path.exists(path):
        rep.add("2 tracker", False, f"absent: {path}")
        return
    try:
        from openpyxl import load_workbook
    except ImportError:
        rep.add("2 tracker", False, "openpyxl is missing, so the tracker was not read")
        return
    ws = load_workbook(path)["Tasks"]

    def val(r, c):
        v = ws.cell(row=r, column=c).value
        return "" if v is None else str(v).strip()

    rows = [r for r in range(2, ws.max_row + 1)
            if val(r, ID_COL) or val(r, NAME_COL)]
    if not rows:
        rep.add("2 tracker", False, f"{path} has no task rows")
        return
    closed_blank = [r for r in rows
                    if val(r, STATUS_COL) == "Completed" and not val(r, EVIDENCE_COL)]
    measured = [r for r in rows if all(val(r, c) for c in MEASURED.values())]
    if closed_blank:
        rep.add("2 tracker", False,
                f"{len(closed_blank)} closed row(s) with an empty evidence cell: "
                + ", ".join(f"row {r}" for r in closed_blank[:6]))
    elif not measured:
        rep.add("2 tracker", False,
                f"{len(rows)} row(s), none with harness, model, tokens and review rounds")
    else:
        rep.add("2 tracker", True,
                f"{len(rows)} row(s), {len(measured)} fully measured, "
                f"0 closed without evidence")


def check_stop_log(root, rep):
    path = os.path.join(root, "evidence", "day-04", "decisions.json")
    if not os.path.exists(path):
        rep.add("3 stop log", False, f"absent: {path}")
        return
    try:
        data = json.load(open(path, encoding="utf-8"))
    except json.JSONDecodeError as exc:
        rep.add("3 stop log", False, f"{path} is not valid JSON: {exc}")
        return
    stops = data.get("stops", [])
    if not stops:
        rep.add("3 stop log", False, f"{path} records no stop at all")
        return
    unmarked = [s for s in stops if s.get("kind") not in ("decision", "lookup")]
    homeless = [s for s in stops
                if s.get("kind") == "lookup" and
                (not s.get("belongs_in") or s["belongs_in"].startswith("UNNAMED"))]
    d = sum(1 for s in stops if s.get("kind") == "decision")
    share = d / len(stops) * 100
    if unmarked:
        rep.add("3 stop log", False, f"{len(unmarked)} stop(s) with no decision mark")
    elif homeless:
        rep.add("3 stop log", False,
                f"{len(homeless)} lookup(s) with no file named to hold the answer")
    else:
        rep.add("3 stop log", True,
                f"{len(stops)} stop(s), {d} decision(s), share {share:.0f}%")


def check_guard(root, rep):
    path = os.path.join(root, "evidence", "day-04", "guard.txt")
    if not os.path.exists(path):
        rep.add("4 guard record", False, f"absent: {path}")
        return
    text = open(path, encoding="utf-8").read()
    # Count refusal LINES, not every mention of "exit 2". A summary line such as
    # "result: 2 refusals with exit 2" is prose about the run, not a third
    # refusal — counting occurrences reported 3 blocks for a 2-block capture.
    # A refusal line has to carry a block marker as well as the exit code.
    blocked = re.compile(r"\bBLOCKED\b|\brefus(?:ed|al)\b|\bdenied\b", re.I)
    exit2 = re.compile(r"\bexit(?:\s+code)?\s*2\b", re.I)
    blocks = sum(1 for ln in text.splitlines() if blocked.search(ln) and exit2.search(ln)
                 and not ln.lstrip().lower().startswith(("result", "summary", "total")))
    allowed = re.search(r"\ballow(?:ed)?\b|\bsucceed(?:ed|s)?\b", text, re.I)
    if blocks < 2:
        rep.add("4 guard record", False,
                f"{blocks} refusal(s) with exit 2 recorded. Two are needed.")
    elif not allowed:
        rep.add("4 guard record", False,
                "no allowed write recorded. A guard that blocks everything is an outage.")
    else:
        rep.add("4 guard record", True,
                f"{blocks} refusal(s) with exit 2 and one allowed write · {path}")


GUARD_CASES = [
    # (capture text, expected block count, expected OK)
    ("1 -> exit 2 : BLOCKED boundary\n2 -> exit 2 : BLOCKED credential\n3 -> allowed\n"
     "result: 2 refusals with exit 2, 1 allowed write\n", 2, True),
    ("1 -> exit 2 : BLOCKED boundary\n3 -> allowed\n", 1, False),
    ("1 -> exit 2 : BLOCKED boundary\n2 -> exit 2 : BLOCKED credential\n", 2, False),
]


def self_test():
    """Cover the guard counter. A prose summary is not a third refusal."""
    import tempfile
    for text, want_blocks, want_ok in GUARD_CASES:
        with tempfile.TemporaryDirectory() as d:
            os.makedirs(os.path.join(d, "evidence", "day-04"))
            with open(os.path.join(d, "evidence", "day-04", "guard.txt"), "w",
                      encoding="utf-8") as f:
                f.write(text)
            rep = Report()
            check_guard(d, rep)
            _, ok, detail = rep.rows[0]
            assert ok is want_ok, f"expected ok={want_ok} for {text!r}, got {detail}"
            # The no-allowed-write message reports no count, so check when present.
            m = re.search(r"(\d+) refusal", detail)
            if m:
                assert int(m.group(1)) == want_blocks, \
                    f"expected {want_blocks} blocks, counted {m.group(1)} — {detail}"
    print("self-test PASS")
    print("  a 2-refusal capture with a prose summary counts 2, not 3")
    print("  a 1-refusal capture is refused")
    print("  a capture with no allowed write is refused")
    return 0


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--root", default=".", help="the folder that holds evidence/")
    ap.add_argument("--file", default="tracker.xlsx", help="the tracker workbook")
    ap.add_argument("--module", default="", help="named in the output only")
    ap.add_argument("--self-test", action="store_true")
    a = ap.parse_args()
    if a.self_test:
        return self_test()

    print(f"Day 4 loop check · root {os.path.abspath(a.root)}"
          + (f" · module {a.module}" if a.module else ""))
    print()
    rep = Report()
    check_loop_record(a.root, rep)
    check_tracker(a.root, a.file, rep)
    check_stop_log(a.root, rep)
    check_guard(a.root, rep)
    return rep.done()


if __name__ == "__main__":
    sys.exit(main())
