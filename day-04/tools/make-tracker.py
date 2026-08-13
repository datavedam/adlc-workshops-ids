#!/usr/bin/env python3
"""
Build a training task tracker.

The real IDSNext tracker is a client file and never travels in this repository.
This builds a workbook with the SAME sheet names, the SAME task columns A-AP and
the SAME status lists, so every command taught on Day 4 runs against a file with
the identical shape. A team can point the same commands at their production file
with no change.

What it deliberately does NOT reproduce: the capacity, leave, utilization and
weekly plan formulas. Those belong to the delivery managers, and the Day 4 write
contract says the agent never touches them. A stub sheet is enough to prove the
agent leaves them alone.

Run:  python3 tools/make-tracker.py [--out tracker.xlsx] [--force]
"""

import argparse
import os
import sys

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("openpyxl is missing.  Install it with:  pip install openpyxl")

# ── the task row, exactly as the client workbook orders it ──────────────────
# A-V are theirs. W-AC are the monthly hour spread. AD-AE are hidden filter
# helpers. AF-AP are the ADLC evidence layer, APPENDED so nothing of theirs
# moved and no formula of theirs would need a change.
TASK_COLUMNS = [
    "Task ID", "Case Number", "Customer Name", "Task Name", "Project",
    "Main Module", "Sub Module", "Developer", "Work Type", "Priority",
    "Status", "Start Date", "End Date", "Est Hrs", "Actual Hrs", "% Complete",
    "Duration (d)", "Sprint (Start)", "Cross-Month?", "Cross-Module?",
    "Customer Committed?", "Remaining Hrs",
    "2026-06", "2026-07", "2026-08", "2026-09", "2026-10", "2026-11", "2026-12",
    "Filter Match", "Filter Rank",
    "AI Mode", "Harness", "Model", "Tokens", "Est Cost (USD)", "Review Rounds",
    "Defects (review)", "Defects (escaped)", "EARS spec?",
    "Evidence (the check that proves it)", "Tollgate",
]

FIRST_ADLC_COL = TASK_COLUMNS.index("AI Mode") + 1          # 32 -> AF
EVIDENCE_COL = TASK_COLUMNS.index("Evidence (the check that proves it)") + 1
STATUS_COL = TASK_COLUMNS.index("Status") + 1
OWNER_COL = TASK_COLUMNS.index("Developer") + 1
EST_COL = TASK_COLUMNS.index("Est Hrs") + 1
NAME_COL = TASK_COLUMNS.index("Task Name") + 1
ID_COL = 1

STATUSES = ["Not Started", "In Progress", "Clarification", "On Hold",
            "QA Pending", "QA In Progress", "Completed"]
WORK_TYPES = ["New Requirement", "Production Issue", "Tech Upgrade",
              "Internal Task", "Internal Issue"]
PRIORITIES = ["Critical", "High", "Medium", "Low"]
AI_MODES = ["unassisted", "chat", "agentic"]
STATUS_WORDS = ["OBSERVED", "DERIVED", "PROPOSED", "APPROVED", "OPEN"]

HEAD = PatternFill("solid", fgColor="EFEBF9")
ADLC = PatternFill("solid", fgColor="E4F0DC")
BOLD = Font(bold=True, color="3A2B57")
ROWS = 301


def _header(ws, labels, adlc_from=None):
    for i, label in enumerate(labels, start=1):
        c = ws.cell(row=1, column=i, value=label)
        c.font = BOLD
        c.fill = ADLC if adlc_from and i >= adlc_from else HEAD
        c.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = max(11, min(34, len(label) + 3))
    ws.freeze_panes = "A2"


def build(path):
    """Write the workbook and return its path."""
    wb = Workbook()

    ws = wb.active
    assert ws is not None, "a new workbook always has one sheet"
    ws.title = "Tasks"
    _header(ws, TASK_COLUMNS, adlc_from=FIRST_ADLC_COL)
    ws.cell(row=ROWS, column=1, value=None)          # reserve the row range

    lists = wb.create_sheet("Lists")
    _header(lists, ["Status", "Priority", "Work Type", "AI Mode", "Status word"])
    for i in range(max(len(STATUSES), len(WORK_TYPES), len(STATUS_WORDS))):
        for col, seq in enumerate((STATUSES, PRIORITIES, WORK_TYPES, AI_MODES,
                                   STATUS_WORDS), start=1):
            if i < len(seq):
                lists.cell(row=i + 2, column=col, value=seq[i])

    team = wb.create_sheet("Team")
    _header(team, ["Dev ID", "Name", "Role", "Home Module", "Base Capacity (hrs/mo)"])

    for name in ("Availability", "Utilization", "Timeline"):
        stub = wb.create_sheet(name)
        _header(stub, ["Owned by the delivery manager. The agent reads this sheet only."])

    tg = wb.create_sheet("Tollgates")
    _header(tg, ["Tollgate", "Day", "Pass criterion (what must be SHOWN)",
                 "Evidence — path or Task ID", "Verified live?",
                 "Signed by (IDSNext)", "Date", "Status"])
    for i, (gate, day, crit) in enumerate([
        ("TG4", 5, "The loop ran end to end on one requirement, with a path for each stage"),
        ("TG4", 5, "Tracker rows carry harness, model, tokens, review rounds and defects"),
        ("TG4", 5, "No closed task row has an empty evidence cell"),
        ("TG4", 5, "The stop log shows the count and the decision share"),
        ("TG4", 5, "The guard refused both violations with exit code 2, and allowed a normal write"),
    ], start=2):
        tg.cell(row=i, column=1, value=gate)
        tg.cell(row=i, column=2, value=day)
        tg.cell(row=i, column=3, value=crit)
        tg.cell(row=i, column=8, value="Not started")

    gates = wb.create_sheet("Gates")
    _header(gates, ["ID", "Day", "Gate", "What it is for", "Blocking?",
                    "PROOF — the bad change it stopped", "Task ID"])
    gates.cell(row=2, column=1, value="G-04-1")
    gates.cell(row=2, column=3, value="PreToolUse guard — repository boundary")
    gates.cell(row=3, column=1, value="G-04-2")
    gates.cell(row=3, column=3, value="PreToolUse guard — credential in content")
    gates.cell(row=4, column=1, value="G-04-3")
    gates.cell(row=4, column=3, value="PreToolUse guard — tracker formula columns")

    adrs = wb.create_sheet("ADRs")
    _header(adrs, ["ADR ID", "Day", "Module", "Decision", "Context",
                   "Consequences", "Strongest counter", "Rebuttal or revision",
                   "Status word", "Owner", "Date"])

    risks = wb.create_sheet("Risks")
    _header(risks, ["ID", "Raised (day)", "Description", "Task ID", "Severity",
                    "Owner", "Due date", "Status"])

    readme = wb.create_sheet("Read Me", 0)
    _header(readme, ["Rules of evidence for this workbook"])
    for i, line in enumerate([
        "A task reaches Completed only when the evidence cell names the check that proves it.",
        "Evidence is a command and its output path, or a test name that fails without the change.",
        "Done, tested locally, and works as expected are not evidence.",
        "The agent may write columns A to P on a proposed row, and AF to AP on its own rows.",
        "The agent may append a row to ADRs, Risks and Gates.",
        "The agent may never write columns Q to AE, the Team, Availability, Utilization or Timeline sheets.",
        "The agent may never write a tollgate signature. A named person signs.",
        "A proposed value stays PROPOSED until a named person approves it.",
        "Self-reported speed is not evidence. A number with no meter reading behind it is an opinion.",
    ], start=2):
        readme.cell(row=i, column=1, value=line)

    wb.save(path)
    return path


def main():
    ap = argparse.ArgumentParser(description="Build a training task tracker.")
    ap.add_argument("--out", default="tracker.xlsx")
    ap.add_argument("--force", action="store_true", help="overwrite an existing file")
    a = ap.parse_args()
    if os.path.exists(a.out) and not a.force:
        sys.exit(f"{a.out} exists already. Use --force to overwrite it.")
    parent = os.path.dirname(os.path.abspath(a.out))
    os.makedirs(parent, exist_ok=True)
    build(a.out)
    print(f"wrote {a.out}")
    print(f"  Tasks columns A-{get_column_letter(len(TASK_COLUMNS))} "
          f"· ADLC layer starts at {get_column_letter(FIRST_ADLC_COL)}")
    print("  sheets: " + ", ".join(["Read Me", "Tasks", "Lists", "Team",
                                    "Availability", "Utilization", "Timeline",
                                    "Tollgates", "Gates", "ADRs", "Risks"]))
    return 0


if __name__ == "__main__":
    sys.exit(main())
