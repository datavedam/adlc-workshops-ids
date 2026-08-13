#!/usr/bin/env python3
"""
Read and write the task tracker inside the Day 4 loop.

Four commands, and the write contract is enforced in code:

  show     print the rows that are still open
  propose  turn EARS criteria in a SPEC.md into proposed task rows
  gate     print only the rows that need a human decision
  check    fail when a row claims Completed with an empty evidence cell

The tool writes columns A-P on a row it created, and AF-AP on its own rows.
It never writes Q-AE (formulas and filter helpers), never writes the Team,
Availability, Utilization or Timeline sheets, and never writes a signature.

Run:  python3 tools/tracker.py <command> --file tracker.xlsx [...]
      python3 tools/tracker.py --self-test
"""

import argparse
import importlib.util
import os
import re
import sys
import tempfile

try:
    from openpyxl import load_workbook
except ImportError:
    sys.exit("openpyxl is missing.  Install it with:  pip install openpyxl")

HERE = os.path.dirname(os.path.abspath(__file__))

# Column letters, so a reader can check them against the sheet in front of them.
COL = {"id": 1, "name": 4, "project": 5, "module": 6, "owner": 8,
       "work_type": 9, "priority": 10, "status": 11, "est": 14, "pct": 16,
       "committed": 21,
       "ai_mode": 32, "harness": 33, "model": 34, "tokens": 35, "cost": 36,
       "rounds": 37, "def_review": 38, "def_escaped": 39, "ears": 40,
       "evidence": 41, "tollgate": 42}
LOCKED = range(17, 32)          # Q..AE — formulas and hidden filter helpers
READ_ONLY_SHEETS = ("Team", "Availability", "Utilization", "Timeline")
OPEN_STATUSES = ("Not Started", "In Progress", "Clarification", "On Hold",
                 "QA Pending", "QA In Progress")

# EARS criteria carry SHALL. That single word is a better detector than any
# heading pattern, because every EARS shape contains it and prose rarely does.
EARS = re.compile(r"\bSHALL\b")
BULLET = re.compile(r"^\s*(?:[-*+]\s+|\d+[.)]\s+)")


def _load(path):
    if not os.path.exists(path):
        sys.exit(f"no workbook at {path}. Build one with tools/make-tracker.py")
    wb = load_workbook(path)
    if "Tasks" not in wb.sheetnames:
        sys.exit(f"{path} has no Tasks sheet, so it is not a task tracker")
    return wb


def _rows(ws):
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=COL["id"]).value or \
           ws.cell(row=r, column=COL["name"]).value:
            yield r


def _val(ws, r, key):
    v = ws.cell(row=r, column=COL[key]).value
    return "" if v is None else str(v).strip()


def _first_free_row(ws):
    used = list(_rows(ws))
    return (max(used) + 1) if used else 2


def cmd_show(a):
    ws = _load(a.file)["Tasks"]
    n = 0
    print(f"{'ROW':>4}  {'TASK':<10} {'STATUS':<15} {'OWNER':<14} NAME")
    for r in _rows(ws):
        status = _val(ws, r, "status") or "(no status)"
        if a.all or status in OPEN_STATUSES or not status:
            n += 1
            print(f"{r:>4}  {_val(ws, r, 'id'):<10} {status:<15} "
                  f"{_val(ws, r, 'owner') or '(none)':<14} {_val(ws, r, 'name')[:52]}")
    print(f"\n{n} row(s) shown.")
    return 0


def read_criteria(spec_path):
    """Return the EARS criteria in a SPEC.md, in file order."""
    if not os.path.exists(spec_path):
        sys.exit(f"no specification at {spec_path}")
    out = []
    for line in open(spec_path, encoding="utf-8"):
        text = BULLET.sub("", line).strip().rstrip("|").strip()
        if EARS.search(text) and len(text.split()) >= 4:
            out.append(re.sub(r"\s+", " ", text))
    return out


def cmd_propose(a):
    criteria = read_criteria(a.spec)
    if not criteria:
        sys.exit(f"{a.spec} holds no EARS criterion. A criterion contains SHALL.")
    wb = _load(a.file)
    ws = wb["Tasks"]
    start = _first_free_row(ws)
    prefix = a.prefix
    for i, text in enumerate(criteria):
        r = start + i
        ws.cell(row=r, column=COL["id"], value=f"{prefix}-{i + 1:02d}")
        ws.cell(row=r, column=COL["name"], value=text[:180])
        ws.cell(row=r, column=COL["project"], value=a.project)
        ws.cell(row=r, column=COL["module"], value=a.module)
        ws.cell(row=r, column=COL["owner"], value=a.owner)
        ws.cell(row=r, column=COL["work_type"], value=a.work_type)
        ws.cell(row=r, column=COL["priority"], value="Medium")
        ws.cell(row=r, column=COL["status"], value="Not Started")
        ws.cell(row=r, column=COL["est"], value="PROPOSED")
        ws.cell(row=r, column=COL["ai_mode"], value="agentic")
        ws.cell(row=r, column=COL["ears"], value="Yes")
        ws.cell(row=r, column=COL["tollgate"], value=a.tollgate)
    wb.save(a.file)
    print(f"proposed {len(criteria)} row(s) into {a.file}, rows "
          f"{start} to {start + len(criteria) - 1}")
    print("every estimate is PROPOSED. A lead corrects it, and that correction "
          "is a decision.")
    return 0


def gate_rows(ws):
    """Rows that cannot move without a human. Each one is a real decision."""
    out = []
    for r in _rows(ws):
        status, why = _val(ws, r, "status"), []
        if _val(ws, r, "est").upper() == "PROPOSED":
            why.append("estimate needs a lead")
        if not _val(ws, r, "owner"):
            why.append("no owner")
        if status == "Clarification":
            why.append("waits on an answer")
        if status == "Completed" and not _val(ws, r, "evidence"):
            why.append("closed with no evidence")
        if why:
            out.append((r, _val(ws, r, "id"), "; ".join(why)))
    return out


def cmd_gate(a):
    ws = _load(a.file)["Tasks"]
    rows = gate_rows(ws)
    for r, tid, why in rows:
        print(f"{r:>4}  {tid:<10} {why}")
    print(f"\n{len(rows)} row(s) need a person.")
    return 0


def check_rows(ws):
    """Closed rows with no evidence. The one rule the whole day turns on."""
    return [(r, _val(ws, r, "id"), _val(ws, r, "name")[:60])
            for r in _rows(ws)
            if _val(ws, r, "status") == "Completed" and not _val(ws, r, "evidence")]


def cmd_check(a):
    ws = _load(a.file)["Tasks"]
    bad = check_rows(ws)
    if not bad:
        print("PASS — every closed row names the check that proves it.")
        return 0
    print(f"FAIL — {len(bad)} closed row(s) have an empty evidence cell:")
    for r, tid, name in bad:
        print(f"    row {r:>4}  {tid:<10} {name}")
    print("\nWrite the command and its output path in column AO, or reopen the row.")
    return 1


def self_test():
    """Build a workbook, propose from a sample specification, prove both gates."""
    spec_src = importlib.util.spec_from_file_location(
        "mk", os.path.join(HERE, "make-tracker.py"))
    assert spec_src and spec_src.loader, "make-tracker.py must sit beside this file"
    mk = importlib.util.module_from_spec(spec_src)
    spec_src.loader.exec_module(mk)

    with tempfile.TemporaryDirectory() as d:
        book, spec = os.path.join(d, "t.xlsx"), os.path.join(d, "SPEC.md")
        mk.build(book)
        open(spec, "w", encoding="utf-8").write(
            "# Spec\n\n## Acceptance criteria\n\n"
            "- WHEN a user opens the card the module SHALL show the month total.\n"
            "- IF a source value is absent the module SHALL show an open marker.\n"
            "- The module SHALL reconcile the department lines to the headline total.\n"
            "\nThis sentence has no criterion in it at all.\n")

        found = read_criteria(spec)
        assert len(found) == 3, f"expected 3 criteria, found {len(found)}"

        args = argparse.Namespace(file=book, spec=spec, owner="A. Lead",
                                  project="FX1", module="overview",
                                  work_type="New Requirement", prefix="D4",
                                  tollgate="TG4")
        cmd_propose(args)

        wb = _load(book)
        ws = wb["Tasks"]
        assert len(list(_rows(ws))) == 3, "propose did not write three rows"
        assert not check_rows(ws), "a Not Started row must not fail the check"
        assert len(gate_rows(ws)) == 3, "every proposed estimate needs a lead"

        for c in LOCKED:                       # the write contract, verified
            assert ws.cell(row=2, column=c).value is None, \
                f"column {c} is locked and must stay empty"
        for name in READ_ONLY_SHEETS:
            assert wb[name].max_row == 1, f"{name} must stay untouched"

        ws.cell(row=2, column=COL["status"], value="Completed")
        wb.save(book)
        ws = _load(book)["Tasks"]
        assert check_rows(ws), "a closed row with no evidence must fail"

        wb = _load(book)
        wb["Tasks"].cell(row=2, column=COL["evidence"],
                         value="python3 -m pytest tests/test_total.py -k month · "
                               "evidence/day-04/test-total.txt")
        wb.save(book)
        assert not check_rows(_load(book)["Tasks"]), "evidence must clear the check"

    print("self-test PASS")
    print("  read 3 EARS criteria and skipped 1 prose line")
    print("  wrote 3 proposed rows, all of them held for a lead")
    print("  columns Q to AE stayed empty, read-only sheets stayed untouched")
    print("  a closed row failed with no evidence and passed with evidence")
    return 0


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--self-test", action="store_true")
    sub = ap.add_subparsers(dest="cmd")

    p = sub.add_parser("show", help="print the open rows")
    p.add_argument("--file", default="tracker.xlsx")
    p.add_argument("--all", action="store_true", help="closed rows as well")
    p.set_defaults(fn=cmd_show)

    p = sub.add_parser("propose", help="EARS criteria become proposed task rows")
    p.add_argument("--file", default="tracker.xlsx")
    p.add_argument("--spec", required=True)
    p.add_argument("--owner", required=True)
    p.add_argument("--project", default="FX1")
    p.add_argument("--module", default="")
    p.add_argument("--work-type", dest="work_type", default="New Requirement")
    p.add_argument("--prefix", default="D4")
    p.add_argument("--tollgate", default="TG4")
    p.set_defaults(fn=cmd_propose)

    p = sub.add_parser("gate", help="print the rows that need a person")
    p.add_argument("--file", default="tracker.xlsx")
    p.set_defaults(fn=cmd_gate)

    p = sub.add_parser("check", help="closed rows must name their check")
    p.add_argument("--file", default="tracker.xlsx")
    p.set_defaults(fn=cmd_check)

    a = ap.parse_args()
    if a.self_test:
        return self_test()
    if not a.cmd:
        ap.print_help()
        return 2
    return a.fn(a)


if __name__ == "__main__":
    sys.exit(main())
